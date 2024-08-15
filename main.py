import customtkinter as ctk
import openpyxl
from openpyxl import Workbook
from tkinter import messagebox


def Create_xml():
    name = Name_entry.get()
    age = Age_entry.get()
    birthday = Birthday_entry.get()
    nationality = Nationality_entry.get()
    position = Position_entry.get()
    hourly_pay = Hourly_entry.get()

    try:
        workbook = openpyxl.load_workbook('data.xlsx')
    except FileNotFoundError:
        workbook = Workbook()

    sheet = workbook.active

    next_row = sheet.max_row + 1
    sheet[f'A{next_row}'] = name
    sheet[f'B{next_row}'] = age
    sheet[f'C{next_row}'] = birthday
    sheet[f'D{next_row}'] = nationality
    sheet[f'E{next_row}'] = position
    sheet[f'F{next_row}'] = hourly_pay

    workbook.save('data.xlsx')

    Display_data_in_excel_frame()


def Display_data_in_excel_frame():
    for widget in excel_frame.winfo_children():
        widget.destroy()

    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook.active

    headers = ["Name", "Age", "Birthday", "Nationality", "Position", "Hourly Pay"]
    for col_num, header in enumerate(headers, start=1):
        header_label = ctk.CTkLabel(excel_frame, text=header, font=font2)
        header_label.grid(row=0, column=col_num - 1, padx=5, pady=5, sticky='nsew')

    for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_text = f"{row[0]}, {row[1]}, {row[2]}, {row[3]}, {row[4]}, {row[5]}"
        row_button = ctk.CTkButton(excel_frame, text=row_text, command=lambda r=row_num: select_row(r))
        row_button.grid(row=row_num, column=0, columnspan=len(headers), padx=5, pady=5, sticky='nsew')


def select_row(row_num):
    global selected_row
    selected_row = row_num


def remove_selected_row():
    if selected_row is None:
        messagebox.showerror("Error", "No row selected")
        return

    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook.active

    sheet.delete_rows(selected_row)
    workbook.save('data.xlsx')

    Display_data_in_excel_frame()


app = ctk.CTk()
app.title('Data Entry')
app.geometry('1160x600')
app.configure(bg='#1e1e2f')

font1 = ('Arial', 25, 'bold')
font2 = ('Arial', 18, 'bold')

selected_row = None

# Company branding
company_label = ctk.CTkLabel(app, text="EXAMPLE COMPANY", font=('Arial', 35, 'bold'), fg_color='#2e2e3d')
company_label.pack(pady=20)

# Main frame
main_frame = ctk.CTkFrame(app, fg_color='#2e2e3d', bg_color='#1e1e2f', border_color='#4b4b5a', border_width=3)
main_frame.pack(pady=20, padx=20, fill='both', expand=True)

# Left frame for data entry
entry_frame = ctk.CTkFrame(main_frame, fg_color='#2e2e3d', bg_color='#2e2e3d')
entry_frame.grid(row=0, column=0, padx=20, pady=20, sticky='nsew')

entry_title = ctk.CTkLabel(entry_frame, text="Enter Your Details", font=font1, fg_color='#2e2e3d')
entry_title.grid(row=0, column=0, columnspan=2, pady=10)

# Entry fields
Name_entry = ctk.CTkEntry(entry_frame, placeholder_text="Name", width=250)
Name_entry.grid(row=1, column=0, padx=10, pady=10, sticky='ew')

Age_entry = ctk.CTkEntry(entry_frame, placeholder_text="Age", width=250)
Age_entry.grid(row=1, column=1, padx=10, pady=10, sticky='ew')

Birthday_entry = ctk.CTkEntry(entry_frame, placeholder_text="Birthday", width=250)
Birthday_entry.grid(row=2, column=0, padx=10, pady=10, sticky='ew')

Nationality_entry = ctk.CTkEntry(entry_frame, placeholder_text="Nationality", width=250)
Nationality_entry.grid(row=2, column=1, padx=10, pady=10, sticky='ew')

Position_entry = ctk.CTkEntry(entry_frame, placeholder_text="Position", width=250)
Position_entry.grid(row=3, column=0, padx=10, pady=10, sticky='ew')

Hourly_entry = ctk.CTkEntry(entry_frame, placeholder_text="Hourly Pay", width=250)
Hourly_entry.grid(row=3, column=1, padx=10, pady=10, sticky='ew')

# Confirm button
Confirm_button = ctk.CTkButton(entry_frame, text="Save to Excel", command=Create_xml)
Confirm_button.grid(row=4, column=0, columnspan=2, pady=20, sticky="nsew")

# Thin separator
separator = ctk.CTkFrame(main_frame, width=2, fg_color="#4b4b5a")
separator.grid(row=0, column=1, padx=10, pady=20, sticky="ns")

# Right frame for displaying Excel data
excel_frame = ctk.CTkFrame(main_frame, fg_color='#2e2e3d', bg_color='#2e2e3d')
excel_frame.grid(row=0, column=2, padx=20, pady=20, sticky='nsew')

# Row removal button
remove_row_button = ctk.CTkButton(main_frame, text="Remove Selected Row", command=remove_selected_row)
remove_row_button.grid(row=1, column=2, pady=10, padx=20, sticky="ew")

# Configure grid weights
main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_columnconfigure(1, weight=0)
main_frame.grid_columnconfigure(2, weight=2)
main_frame.grid_rowconfigure(0, weight=1)
main_frame.grid_rowconfigure(1, weight=0)

# Display Excel data header initially
Display_data_in_excel_frame()

app.mainloop()
